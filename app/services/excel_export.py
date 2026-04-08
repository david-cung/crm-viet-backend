"""Excel template generator — tạo file .xlsx mẫu cho từng module."""
from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.excel_import import HEADERS

# Màu header
HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

# Dữ liệu mẫu mỗi module
SAMPLE_DATA: dict[str, list[list[Any]]] = {
    "employees": [
        ["NV-001", "Nguyễn Văn An", "1990-05-15", "012345678901", "0901234567",
         "an.nguyen@cty.vn", "Kinh doanh", "Sales Executive", "full_time",
         "2022-01-10", 15000000, 500000, 730000, 300000, 1,
         "Vietcombank", "1234567890", "CN Hà Nội"],
        ["NV-002", "Trần Thị Bình", "1995-08-20", "098765432100", "0912345678",
         "binh.tran@cty.vn", "Marketing", "Marketing Specialist", "full_time",
         "2023-03-01", 12000000, 500000, 730000, 0, 0,
         "Techcombank", "9876543210", "CN TP.HCM"],
    ],
    "payroll": [
        ["NV-001", 4, 2026, 26, 25.5, 4, 1000000, 0, "Tháng 4/2026"],
        ["NV-002", 4, 2026, 26, 26, 0, 0, 0, ""],
    ],
    "commission": [
        ["NV-001", 4, 2026, 85000000, "Đạt 85% target"],
        ["NV-002", 4, 2026, 120000000, "Vượt target 20%"],
    ],
    "cashflow": [
        ["2026-04-01", "thu", "Doanh thu bán hàng", 50000000, "Hóa đơn #001", "Nguyễn Văn An"],
        ["2026-04-02", "chi", "Chi phí văn phòng", 2500000, "Thuê văn phòng T4", "Admin"],
        ["2026-04-03", "chi", "Lương nhân viên", 80000000, "Lương tháng 3", "Kế toán"],
    ],
    "products": [
        ["SP-001", "Phần mềm CRM Basic", "Software", "license", 5000000, 8000000, 50, 10, ""],
        ["SP-002", "Phần mềm CRM Pro", "Software", "license", 10000000, 15000000, 30, 5, "Cao cấp"],
        ["SP-003", "Tư vấn triển khai", "Service", "giờ", 500000, 1200000, 1000, 100, "Dịch vụ"],
    ],
    "debts": [
        ["Công ty ABC", "thu", "HD-2026-001", "2026-01-15", 120000000, 50000000, "2026-04-15", 0.5, ""],
        ["Nhà cung cấp XYZ", "chi", "PO-2026-010", "2026-02-01", 45000000, 45000000, "2026-03-01", 0, "Đã thanh toán"],
    ],
}

NOTES: dict[str, str] = {
    "employees": "Loại HĐ: full_time | part_time | probation | intern",
    "payroll": "Tháng: 1-12, Năm: 4 chữ số. Giờ OT không bắt buộc.",
    "commission": "Doanh số tính bằng VND, không cần ký tự phân cách.",
    "cashflow": "Loại: thu hoặc chi. Ngày định dạng yyyy-mm-dd hoặc dd/mm/yyyy.",
    "products": "Đơn vị: cái, kg, hộp, license, giờ... Giá nhập bằng VND.",
    "debts": "Loại: thu (phải thu) hoặc chi (phải trả). Lãi phạt tính %/tháng.",
}


def generate_template(module: str) -> bytes:
    headers = HEADERS.get(module)
    if not headers:
        raise ValueError(f"Module không hỗ trợ: {module}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = module.capitalize()

    # Header row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # Sample data rows
    for row_idx, sample_row in enumerate(SAMPLE_DATA.get(module, []), start=2):
        for col_idx, val in enumerate(sample_row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Note row
    note = NOTES.get(module, "")
    if note:
        note_row = ws.max_row + 2
        note_cell = ws.cell(row=note_row, column=1, value=f"Ghi chú: {note}")
        note_cell.font = Font(italic=True, color="888888")
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(headers))

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
