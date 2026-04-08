"""
Excel import helpers — đọc file .xlsx, validate từng hàng, trả kết quả chi tiết.
Hỗ trợ: employees, payroll, commission, cashflow, products, debts.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

# ---------------------------------------------------------------------------
# TEMPLATE HEADERS (phải khớp chính xác hàng 1 của file Excel)
# ---------------------------------------------------------------------------

HEADERS: dict[str, list[str]] = {
    "employees": [
        "Mã NV", "Họ tên", "Ngày sinh", "CCCD", "SĐT", "Email",
        "Phòng ban", "Chức vụ", "Loại HĐ", "Ngày vào",
        "Lương cơ bản", "PC đi lại", "PC ăn", "PC điện thoại",
        "Số NP phụ thuộc", "Tên NH", "Số TK", "Chi nhánh NH",
    ],
    "payroll": [
        "Mã NV", "Tháng", "Năm",
        "Ngày công chuẩn", "Ngày công thực", "Giờ OT",
        "Thưởng thêm", "Khấu trừ khác", "Ghi chú",
    ],
    "commission": [
        "Mã NV", "Tháng", "Năm", "Doanh số cá nhân (VND)", "Ghi chú",
    ],
    "cashflow": [
        "Ngày", "Loại", "Danh mục", "Số tiền", "Ghi chú", "Người tạo",
    ],
    "products": [
        "SKU", "Tên sản phẩm", "Danh mục", "Đơn vị",
        "Giá vốn", "Giá bán", "Tồn kho hiện tại", "Tồn tối thiểu", "Ghi chú",
    ],
    "debts": [
        "Tên khách hàng", "Loại", "Số HĐ", "Ngày HĐ",
        "Tổng tiền", "Đã trả", "Hạn TT", "Lãi phạt (%/tháng)", "Ghi chú",
    ],
}


# ---------------------------------------------------------------------------
# DATA CLASSES
# ---------------------------------------------------------------------------

@dataclass
class RowError:
    row: int
    column: str
    error: str


@dataclass
class ImportResult:
    module: str
    total_rows: int = 0
    success_rows: int = 0
    failed_rows: int = 0
    errors: list[RowError] = field(default_factory=list)
    rows: list[dict[str, Any]] = field(default_factory=list)  # validated rows

    def to_dict(self) -> dict:
        return {
            "module": self.module,
            "totalRows": self.total_rows,
            "successRows": self.success_rows,
            "failedRows": self.failed_rows,
            "errors": [{"row": e.row, "column": e.column, "error": e.error} for e in self.errors],
        }


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _cell_decimal(val: Any, col: str, row: int, errors: list[RowError]) -> Decimal | None:
    s = _cell_str(val)
    if not s:
        return None
    try:
        return Decimal(s.replace(",", "").replace(".", "."))
    except InvalidOperation:
        errors.append(RowError(row, col, f"'{s}' không phải số hợp lệ"))
        return None


def _cell_int(val: Any, col: str, row: int, errors: list[RowError]) -> int | None:
    s = _cell_str(val)
    if not s:
        return None
    try:
        return int(float(s.replace(",", "")))
    except (ValueError, TypeError):
        errors.append(RowError(row, col, f"'{s}' phải là số nguyên"))
        return None


def _cell_date(val: Any, col: str, row: int, errors: list[RowError]) -> str | None:
    """Return ISO date string (YYYY-MM-DD) or None."""
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.strftime("%Y-%m-%d")
    s = _cell_str(val)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    errors.append(RowError(row, col, f"'{s}' không đúng định dạng ngày (dd/mm/yyyy hoặc yyyy-mm-dd)"))
    return None


def _require_str(val: Any, col: str, row: int, errors: list[RowError]) -> str | None:
    s = _cell_str(val)
    if not s:
        errors.append(RowError(row, col, "Trường bắt buộc, không được để trống"))
        return None
    return s


def _require_decimal(val: Any, col: str, row: int, errors: list[RowError]) -> Decimal | None:
    d = _cell_decimal(val, col, row, errors)
    if d is None and not errors:
        errors.append(RowError(row, col, "Trường bắt buộc, không được để trống"))
    return d


# ---------------------------------------------------------------------------
# WORKBOOK LOADER
# ---------------------------------------------------------------------------

def _load_wb(file_bytes: bytes) -> openpyxl.Workbook:
    try:
        return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except InvalidFileException as e:
        raise ValueError(f"File không hợp lệ hoặc không phải .xlsx: {e}") from e


def _validate_headers(ws, expected: list[str]) -> list[str] | None:
    """Return list of actual headers or None if row 1 is empty."""
    row1 = [_cell_str(ws.cell(1, c).value) for c in range(1, len(expected) + 1)]
    if all(h == "" for h in row1):
        return None
    return row1


def _check_header_match(actual: list[str], expected: list[str]) -> list[str]:
    """Return list of error messages for mismatched headers."""
    errs = []
    for i, exp in enumerate(expected):
        act = actual[i] if i < len(actual) else ""
        if act != exp:
            errs.append(f"Cột {i+1}: tìm thấy '{act}', cần '{exp}'")
    return errs


# ---------------------------------------------------------------------------
# MODULE PARSERS
# ---------------------------------------------------------------------------

def parse_employees(file_bytes: bytes) -> ImportResult:
    res = ImportResult(module="employees")
    wb = _load_wb(file_bytes)
    ws = wb.active
    headers = _validate_headers(ws, HEADERS["employees"])
    if headers is None:
        raise ValueError("File trống hoặc không đúng template employees")
    hdr_errors = _check_header_match(headers, HEADERS["employees"])
    if hdr_errors:
        raise ValueError("Header không đúng template:\n" + "\n".join(hdr_errors))

    for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        res.total_rows += 1
        row_errors: list[RowError] = []
        h = HEADERS["employees"]

        employee_code = _require_str(row[0], h[0], ri, row_errors)
        full_name = _require_str(row[1], h[1], ri, row_errors)
        date_of_birth = _cell_date(row[2], h[2], ri, row_errors)
        id_card = _cell_str(row[3])
        phone = _cell_str(row[4])
        personal_email = _cell_str(row[5])
        department = _cell_str(row[6])
        position = _cell_str(row[7])
        contract_type = _cell_str(row[8]) or "full_time"
        start_date = _cell_date(row[9], h[9], ri, row_errors)
        base_salary = _cell_decimal(row[10], h[10], ri, row_errors) or Decimal(0)
        allow_transport = _cell_decimal(row[11], h[11], ri, row_errors) or Decimal(0)
        allow_meal = _cell_decimal(row[12], h[12], ri, row_errors) or Decimal(0)
        allow_phone = _cell_decimal(row[13], h[13], ri, row_errors) or Decimal(0)
        dependent_count = _cell_int(row[14], h[14], ri, row_errors) or 0
        bank_name = _cell_str(row[15])
        bank_account = _cell_str(row[16])
        bank_branch = _cell_str(row[17])

        if row_errors:
            res.errors.extend(row_errors)
            res.failed_rows += 1
        else:
            res.success_rows += 1
            res.rows.append({
                "employeeCode": employee_code,
                "fullName": full_name,
                "dateOfBirth": date_of_birth,
                "idCard": id_card,
                "phone": phone,
                "personalEmail": personal_email,
                "department": department,
                "position": position,
                "contractType": contract_type,
                "startDate": start_date,
                "baseSalary": float(base_salary),
                "allowances": {
                    "transport": float(allow_transport),
                    "meal": float(allow_meal),
                    "phone": float(allow_phone),
                },
                "dependentCount": dependent_count,
                "bankName": bank_name,
                "bankAccount": bank_account,
                "bankBranch": bank_branch,
            })
    return res


def parse_payroll(file_bytes: bytes) -> ImportResult:
    res = ImportResult(module="payroll")
    wb = _load_wb(file_bytes)
    ws = wb.active
    headers = _validate_headers(ws, HEADERS["payroll"])
    if headers is None:
        raise ValueError("File trống hoặc không đúng template payroll")
    hdr_errors = _check_header_match(headers, HEADERS["payroll"])
    if hdr_errors:
        raise ValueError("Header không đúng template:\n" + "\n".join(hdr_errors))

    for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        res.total_rows += 1
        row_errors: list[RowError] = []
        h = HEADERS["payroll"]

        employee_code = _require_str(row[0], h[0], ri, row_errors)
        month = _cell_int(row[1], h[1], ri, row_errors)
        year = _cell_int(row[2], h[2], ri, row_errors)
        standard_days = _cell_decimal(row[3], h[3], ri, row_errors) or Decimal("26")
        actual_days = _cell_decimal(row[4], h[4], ri, row_errors) or Decimal("0")
        ot_hours = _cell_decimal(row[5], h[5], ri, row_errors) or Decimal("0")
        bonus = _cell_decimal(row[6], h[6], ri, row_errors) or Decimal("0")
        other_deductions = _cell_decimal(row[7], h[7], ri, row_errors) or Decimal("0")
        note = _cell_str(row[8])

        if month and not (1 <= month <= 12):
            row_errors.append(RowError(ri, h[1], "Tháng phải từ 1-12"))
        if year and not (2000 <= year <= 2100):
            row_errors.append(RowError(ri, h[2], "Năm không hợp lệ"))

        if row_errors:
            res.errors.extend(row_errors)
            res.failed_rows += 1
        else:
            res.success_rows += 1
            res.rows.append({
                "employeeCode": employee_code,
                "month": month,
                "year": year,
                "standardDays": float(standard_days),
                "actualDays": float(actual_days),
                "otHours": float(ot_hours),
                "bonus": float(bonus),
                "otherDeductions": float(other_deductions),
                "note": note,
            })
    return res


def parse_commission(file_bytes: bytes) -> ImportResult:
    res = ImportResult(module="commission")
    wb = _load_wb(file_bytes)
    ws = wb.active
    headers = _validate_headers(ws, HEADERS["commission"])
    if headers is None:
        raise ValueError("File trống hoặc không đúng template commission")
    hdr_errors = _check_header_match(headers, HEADERS["commission"])
    if hdr_errors:
        raise ValueError("Header không đúng template:\n" + "\n".join(hdr_errors))

    for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        res.total_rows += 1
        row_errors: list[RowError] = []
        h = HEADERS["commission"]

        employee_code = _require_str(row[0], h[0], ri, row_errors)
        month = _cell_int(row[1], h[1], ri, row_errors)
        year = _cell_int(row[2], h[2], ri, row_errors)
        revenue = _require_decimal(row[3], h[3], ri, row_errors) or Decimal(0)
        note = _cell_str(row[4])

        if row_errors:
            res.errors.extend(row_errors)
            res.failed_rows += 1
        else:
            res.success_rows += 1
            res.rows.append({
                "employeeCode": employee_code,
                "month": month,
                "year": year,
                "personalRevenue": float(revenue),
                "note": note,
            })
    return res


def parse_cashflow(file_bytes: bytes) -> ImportResult:
    res = ImportResult(module="cashflow")
    wb = _load_wb(file_bytes)
    ws = wb.active
    headers = _validate_headers(ws, HEADERS["cashflow"])
    if headers is None:
        raise ValueError("File trống hoặc không đúng template cashflow")
    hdr_errors = _check_header_match(headers, HEADERS["cashflow"])
    if hdr_errors:
        raise ValueError("Header không đúng template:\n" + "\n".join(hdr_errors))

    for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        res.total_rows += 1
        row_errors: list[RowError] = []
        h = HEADERS["cashflow"]

        txn_date = _cell_date(row[0], h[0], ri, row_errors)
        txn_type = _cell_str(row[1]).lower()
        category = _require_str(row[2], h[2], ri, row_errors)
        amount = _require_decimal(row[3], h[3], ri, row_errors) or Decimal(0)
        note = _cell_str(row[4])
        created_by = _cell_str(row[5])

        if txn_type not in ("thu", "chi", ""):
            row_errors.append(RowError(ri, h[1], "Loại phải là 'thu' hoặc 'chi'"))
        if not txn_date:
            row_errors.append(RowError(ri, h[0], "Ngày là bắt buộc"))

        if row_errors:
            res.errors.extend(row_errors)
            res.failed_rows += 1
        else:
            res.success_rows += 1
            res.rows.append({
                "date": txn_date,
                "type": txn_type or "chi",
                "category": category,
                "amount": float(amount),
                "note": note,
                "createdBy": created_by,
            })
    return res


def parse_products(file_bytes: bytes) -> ImportResult:
    res = ImportResult(module="products")
    wb = _load_wb(file_bytes)
    ws = wb.active
    headers = _validate_headers(ws, HEADERS["products"])
    if headers is None:
        raise ValueError("File trống hoặc không đúng template products")
    hdr_errors = _check_header_match(headers, HEADERS["products"])
    if hdr_errors:
        raise ValueError("Header không đúng template:\n" + "\n".join(hdr_errors))

    for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        res.total_rows += 1
        row_errors: list[RowError] = []
        h = HEADERS["products"]

        sku = _require_str(row[0], h[0], ri, row_errors)
        name = _require_str(row[1], h[1], ri, row_errors)
        category = _cell_str(row[2])
        unit = _cell_str(row[3]) or "cái"
        cost_price = _cell_decimal(row[4], h[4], ri, row_errors) or Decimal(0)
        sale_price = _cell_decimal(row[5], h[5], ri, row_errors) or Decimal(0)
        stock_qty = _cell_decimal(row[6], h[6], ri, row_errors) or Decimal(0)
        min_stock = _cell_decimal(row[7], h[7], ri, row_errors) or Decimal(0)
        note = _cell_str(row[8])

        if row_errors:
            res.errors.extend(row_errors)
            res.failed_rows += 1
        else:
            res.success_rows += 1
            res.rows.append({
                "sku": sku,
                "name": name,
                "category": category,
                "unit": unit,
                "costPrice": float(cost_price),
                "salePrice": float(sale_price),
                "stockQty": float(stock_qty),
                "minStock": float(min_stock),
                "note": note,
            })
    return res


def parse_debts(file_bytes: bytes) -> ImportResult:
    res = ImportResult(module="debts")
    wb = _load_wb(file_bytes)
    ws = wb.active
    headers = _validate_headers(ws, HEADERS["debts"])
    if headers is None:
        raise ValueError("File trống hoặc không đúng template debts")
    hdr_errors = _check_header_match(headers, HEADERS["debts"])
    if hdr_errors:
        raise ValueError("Header không đúng template:\n" + "\n".join(hdr_errors))

    for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        res.total_rows += 1
        row_errors: list[RowError] = []
        h = HEADERS["debts"]

        customer = _require_str(row[0], h[0], ri, row_errors)
        debt_type = _cell_str(row[1]).lower()
        invoice_no = _cell_str(row[2])
        issue_date = _cell_date(row[3], h[3], ri, row_errors)
        original_amount = _require_decimal(row[4], h[4], ri, row_errors) or Decimal(0)
        paid_amount = _cell_decimal(row[5], h[5], ri, row_errors) or Decimal(0)
        due_date = _cell_date(row[6], h[6], ri, row_errors)
        penalty_rate = _cell_decimal(row[7], h[7], ri, row_errors) or Decimal(0)
        note = _cell_str(row[8])

        if debt_type not in ("thu", "chi", "receivable", "payable", ""):
            row_errors.append(RowError(ri, h[1], "Loại phải là 'thu' hoặc 'chi'"))
        type_map = {"thu": "receivable", "chi": "payable", "receivable": "receivable", "payable": "payable"}
        mapped_type = type_map.get(debt_type, "receivable")

        if row_errors:
            res.errors.extend(row_errors)
            res.failed_rows += 1
        else:
            res.success_rows += 1
            res.rows.append({
                "customerName": customer,
                "debtType": mapped_type,
                "invoiceNumber": invoice_no,
                "issueDate": issue_date,
                "originalAmount": float(original_amount),
                "paidAmount": float(paid_amount),
                "remainingAmount": float(original_amount - paid_amount),
                "dueDate": due_date,
                "penaltyRate": float(penalty_rate),
                "note": note,
            })
    return res


# ---------------------------------------------------------------------------
# DISPATCHER
# ---------------------------------------------------------------------------

PARSERS = {
    "employees": parse_employees,
    "payroll": parse_payroll,
    "commission": parse_commission,
    "cashflow": parse_cashflow,
    "products": parse_products,
    "debts": parse_debts,
}


def parse_excel(module: str, file_bytes: bytes) -> ImportResult:
    if module not in PARSERS:
        raise ValueError(f"Module không hỗ trợ: {module}")
    return PARSERS[module](file_bytes)
